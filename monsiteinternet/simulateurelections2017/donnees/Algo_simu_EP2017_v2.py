# Algorithme en python de résultat aux élections présidentielles
# Note : Aucune case explicitée dans la feuille excel ne doit
# avoir de dépendances avec une case qui continue une formule 
# (point à améliorer)

# Comment ça marche ?
# 1) On importe notre feuille excel qu'on a construite wb()
# 2) On entre les données dans les cases concernées dans la feuille excel (cases bleues) 
# 3) On évalue toute la feuille
# 4) On extrait les données qu'on souhaite !
from openpyxl import load_workbook
import re
import string
import pickle

def convert_wb(string):
	""" 
	Convertit les cases excel en leur valeur.
	Ex : "=round(1-A1,0)" --> "1-sheet["A1"].value"
	"""
	string=string.replace("$","")
	string=convert_round(string)
	string=string.replace("=","") #enlève le = au début de l'expression
	string=re.sub(r"(?P<case>[A-Z]+[0-9]+)",r"sheet['\g<case>'].value",string)

	return string
def convert_round(string):
	"""
	Permet à une case d'activer la fonction round
	"""
	return string.replace(";",",").replace("ROUND","round")


# def convert_simple_variable(string):
# 	"""
# 	Permet de simplifier l'expression avec les cases les plus simples
# 	"""
# 	liste_LettreChiffre=re.findall(r"(?P<case>[A-Z]+[0-9]+)",string)
# 	for lc in liste_LettreChiffre:	#utilisation de parcours de graphe
# 		stack=[lc,[lc]]
# 		while stack:
# 			(lc_courant,path)=stack.pop()
# 			valeur=sheet[lc_courant].value
# 			if re.findall(r"(?P<case>[A-Z]+[0-9]+)",valeur):
# 				stack.append(lc_courant,path + [lc_courant])
# 			else:

# 				for lc in re.findall(r"(?P<case>[A-Z]+[0-9]+)",valeur):

# 				valeur=re.findall(r"(?P<case>[A-Z]+[0-9]+)",valeur)
		

# 		re.findall(r"(?P<case>[A-Z]+[0-9]+)",valeur)

def evaluer_case(sheet,case):
	"""
	Evalue une case d'un fichier wb
	"""
	valeur_conv=convert_wb(sheet[case].value)
	return eval(valeur_conv)

def evaluer_sheet(sheet):
	for lettre in list(string.ascii_uppercase):
		for chiffre in range (1,100):
			case=str(lettre+str(chiffre))
			if "=" in str(sheet[case].value):
				sheet[case].value=evaluer_case(sheet,case)
	return sheet


#récupérer la wb dans notre fichier compressé (cf tuto) sinon ! C'est sûrement plus rapide

# On insère les données entrées par l'utilisateur

donnees_a_maj={"H4":0.01}
def MAJ_donnees_sheet(donnees_a_maj,sheet):
	"""
	MAJ les donnees reçues sous forme de dictionnaire {case}:valeur
	"""
	for key,value in donnees_a_maj.items():
		sheet[key].value=value
	return sheet

def extraire_resultat(sheet):
	"""
	Renvoie un dictionnaire avec
	{Macron:%votes,LePen:%votes,Abstention:Nb}
	"""
	dic_result={}
	case_pv_mac="J29"
	case_pv_lp="J28"
	pv_mac=round(round(sheet[case_pv_mac].value,4)*100,2)
	pv_lp=round(round(sheet[case_pv_lp].value,4)*100,2)
	dic_result["pv_mac"]=pv_mac
	dic_result["pv_lp"]=pv_lp
	return dic_result


def get_resultat_simu(donnees):
	"""
	Renvoie un dictionnaire des résultats. En entrée, les données sont mises sous la forme d'un dico {"CASE":valeur_case}
	"""
	recharger=False
	if recharger:
		wb=load_workbook("Simulateur_deuxieme_tour_EP2017FR.xlsx")
	else: #depickler
		with open('workbook_pickle', 'rb') as fichier:
			mon_depickler = pickle.Unpickler(fichier)
			wb = mon_depickler.load() # on récupère les objets
	sheet= wb.active
	sheet=MAJ_donnees_sheet(donnees,sheet)
	sheet_eval=evaluer_sheet(sheet)
	dic_result=extraire_resultat(sheet)
	return dic_result

if __name__=="__main__":
	#on exécute ce programme tout seul, alors on recharge notre workbook
	workbook=load_workbook("Simulateur_deuxieme_tour_EP2017FR.xlsx")
	with open('workbook_pickle', 'wb') as fichier:
		mon_pickler = pickle.Pickler(fichier) #on crée instance mon_pickler
		# enregistrement ...
		mon_pickler.dump(workbook) #on a enregistré nos objets
	val_abs_NDA=0.5
	val_mac_NDA=0.8
	donnees={"H4":val_abs_NDA,"J4":val_mac_NDA}
	d_result=get_resultat_simu(donnees)
	print(d_result)




# val_abs_NDA=1
# val_mac_NDA=0.8
# donnees={"H4":val_abs_NDA,"J4":val_mac_NDA}
# d_result=get_resultat_simu(donnees)
# print(d_result)
# Il faudra mettre la feuille excel d'origine à disposition !
